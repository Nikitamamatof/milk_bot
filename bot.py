# bot.py
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup

from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os

# TOKEN taken from environment variable (set on Render / Railway)
TOKEN = os.getenv("TOKEN")
if not TOKEN:
    print("Warning: TOKEN not set. Set environment variable TOKEN or Render will fail to start.")

# Products list
PRODUCTS = [
    ("Молоко 2,5% 0,9л", 310),
    ("Молоко 3,2% 0,9л", 340),
    ("Кефир 2,5% 0,9л", 330),
    ("Кефир 1% 0,9л", 330),
    ("Кефир снежок 0,9л", 370),
    ("Сливки кг", 3050),
    ("Сливки 500гр", 1680),
    ("Сметана 15% 200гр", 290),
    ("Сметана 15% 400гр", 460),
    ("Сметана 15% 1кг", 1210),
    ("Сметана 20% 200гр", 310),
    ("Сметана 20% 400гр", 480),
    ("Сметана 20% 600гр", 690),
    ("Сметана 20% 1кг", 1310),
    ("Сулугуни в вакууме кг", 3900),
    ("Сулугуни кг", 3800),
    ("Чечел кг", 4300),
    ("Адыгейский кг", 3000),
    ("Колбасный вакуумный кг", 2900),
    ("Колбасный кг", 2830),
    ("Курт вакуум кг", 3800),
    ("Курт кг", 3700),
    ("Курт ведро 1 л", 2700),
    ("Курт ведро 0,5 л", 1500),
    ("Масло сладко-сливочное кг", 5200),
    ("Масло сладко-сливочное 200гр", 1150),
    ("Масло соленое кг", 5250),
    ("Творог кг", 1600),
    ("Айран турецкий", 135),
    ("Катык 400гр", 1080),
    ("Иримшик кг", 2930),
    ("Иримшик 200 гр", 710),
    ("Жент кг", 2580),
    ("Жент 200 гр", 670),
    ("Жент 500 гр", 1290),
]

EXCHANGE_REQUIRED = {
    "Молоко 2,5% 0,9л",
    "Молоко 3,2% 0,9л",
    "Кефир 2,5% 0,9л",
    "Кефир 1% 0,9л",
    "Сметана 15% 200гр",
    "Сметана 15% 400гр",
    "Сметана 15% 1кг",
    "Сметана 20% 200гр",
    "Сметана 20% 400гр",
    "Сметана 20% 600гр",
    "Сметана 20% 1кг",
    "Айран турецкий",
    "Катык 400гр",
}

class Form(StatesGroup):
    filling = State()

bot = Bot(token=TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())

USERS = {}

HELP_TEXT = (
    "Инструкция:\n"
    "/start_report — начать новый отчёт\n"
    "/cancel — отменить ввод\n\n"
    "Во время ввода:\n"
    "• Введи число (например 12)\n"
    "• Введи 0 — если товара не было\n"
    "• Напиши 'пропустить' или 'skip' — чтобы пропустить позицию\n"
    "После завершения бот пришлёт итог и Excel-файл."
)

@dp.message_handler(commands=[\"start\"])
async def cmd_start(message: types.Message):
    await message.answer(\"Привет! Я бот-отчётчик.\n\" + HELP_TEXT)

@dp.message_handler(commands=[\"help\"])
async def cmd_help(message: types.Message):
    await message.answer(HELP_TEXT)

@dp.message_handler(commands=[\"cancel\"], state=\"*\")
async def cmd_cancel(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    USERS.pop(user_id, None)
    await state.finish()
    await message.answer(\"Ввод отменён. Для начала заново — /start_report\")

@dp.message_handler(commands=[\"start_report\"])
async def cmd_start_report(message: types.Message):
    user_id = message.from_user.id
    USERS[user_id] = {\"idx\": 0, \"step\": \"morning\", \"rows\": {}}
    for name, price in PRODUCTS:
        USERS[user_id][\"rows\"][name] = {\"price\": price, \"morning\": 0.0, \"evening\": 0.0, \"exchange\": 0.0}
    await Form.filling.set()
    name, price = PRODUCTS[0]
    await message.answer(
        f\"Начинаем отчёт.\n\nПозиция 1/{len(PRODUCTS)}:\n{name} — {price} тг\n\"
        f\"Введите количество полученное утром (или 'пропустить').\"
    )

def parse_quantity(text: str):
    text = text.strip().replace(\",\", \".\")
    try:
        return float(text)
    except:
        return None

@dp.message_handler(state=Form.filling)
async def filling_handler(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    if user_id not in USERS:
        await message.answer(\"Сессия не найдена. Напиши /start_report.\")
        await state.finish()
        return

    data = USERS[user_id]
    idx = data[\"idx"]
    step = data[\"step"]
    name, price = PRODUCTS[idx]

    text = message.text.strip().lower()

    # пропуск позиции
    if text in (\"пропустить\", \"skip\"):
        data[\"idx\"] += 1
        data[\"step\"] = \"morning\"
        if data[\"idx\"] >= len(PRODUCTS):
            await state.finish()
            await send_report_and_excel(message, data[\"rows\"])
            USERS.pop(user_id, None)
            return
        next_name, next_price = PRODUCTS[data[\"idx\"]]
        await message.answer(f\"Позиция {data['idx']+1}/{len(PRODUCTS)}:\n{next_name} — {next_price} тг\nВведите количество полученное утром (или 'пропустить').\")
        return

    qty = parse_quantity(text)
    if qty is None:
        await message.answer(\"Введи число (например 5) или 'пропустить'.\")
        return

    if step == \"morning\":
        data[\"rows\"][name][\"morning\"] = qty
        data[\"step\"] = \"evening\"
        await message.answer(f\"{name} — утром {qty}. Теперь введи остаток (вечером).\")
        return
    elif step == \"evening\":
        data[\"rows\"][name][\"evening\"] = qty
        if name in EXCHANGE_REQUIRED:
            data[\"step\"] = \"exchange\"
            await message.answer(f\"{name} — остаток {qty}. Введи обмен (или 0, если не было).\")
            return
        else:
            data[\"idx\"] += 1
            data[\"step\"] = \"morning\"
    elif step == \"exchange\":
        data[\"rows\"][name][\"exchange\"] = qty
        data[\"idx\"] += 1
        data[\"step\"] = \"morning\"

    if data[\"idx\"] >= len(PRODUCTS):
        await state.finish()
        await send_report_and_excel(message, data[\"rows\"])
        USERS.pop(user_id, None)
        return
    next_name, next_price = PRODUCTS[data[\"idx\"]]
    await message.answer(f\"Позиция {data['idx']+1}/{len(PRODUCTS)}:\n{next_name} — {next_price} тг\nВведите количество полученное утром (или 'пропустить').\")

async def send_report_and_excel(message: types.Message, rows: dict):
    total_to_cash = 0.0
    lines = []
    header = [\"Товар\", \"Цена (тг)\", \"Утро\", \"Остаток\", \"Обмен\", \"Продано\", \"Сумма (тг)\"]

    for name, info in rows.items():
        p, m, e, x = info[\"price\"], info[\"morning\"], info[\"evening\"], info[\"exchange\"]
        sold = m - e - x
        amount = sold * p
        total_to_cash += amount
        lines.append((name, p, m, e, x, sold, amount))

    # Markdown table (monospace block) for Telegram
    header_line = f\"| {'Товар':40} | {'Продано':>7} | {'Сумма':>12} |\"
    sep = '|' + '-'*42 + '|' + '-'*9 + '|' + '-'*14 + '|'
    text_lines = []
    text_lines.append(f\"📅 *Отчёт за {datetime.now().strftime('%Y-%m-%d %H:%M')}*\")
    text_lines.append('')
    text_lines.append('```')
    text_lines.append(header_line)
    text_lines.append(sep)
    for name, p, m, e, x, sold, amount in lines:
        sold_display = f\"{sold:.0f}\" if abs(sold - round(sold)) < 1e-6 else f\"{sold:.2f}\"
        amount_display = f\"{amount:,.0f} тг\"
        short_name = (name[:40]).ljust(40)
        text_lines.append(f\"| {short_name} | {sold_display:>7} | {amount_display:>12} |\")
    text_lines.append('```')
    text_lines.append(f\"💰 *Итого к сдаче:* {total_to_cash:,.0f} тг\")

    await message.answer('\\n'.join(text_lines), parse_mode=types.ParseMode.MARKDOWN)

    # Excel creation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = \"Отчёт\"
    ws.merge_cells(\"A1:G1\")
    ws[\"A1\"] = f\"Отчёт торгового представителя — {datetime.now().strftime('%Y-%m-%d %H:%M')}\"
    for col_idx, h in enumerate(header, start=1):
        ws.cell(row=2, column=col_idx, value=h)

    row_num = 3
    for name, p, m, e, x, sold, amount in lines:
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=p)
        ws.cell(row=row_num, column=3, value=m)
        ws.cell(row=row_num, column=4, value=e)
        ws.cell(row=row_num, column=5, value=x)
        ws.cell(row=row_num, column=6, value=sold)
        ws.cell(row=row_num, column=7, value=amount)
        row_num += 1

    ws.cell(row=row_num, column=6, value=\"Итого:\")
    ws.cell(row=row_num, column=7, value=total_to_cash)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value)
            except:
                val = \"\"
            if val is None:
                val = \"\"
            if len(val) > max_length:
                max_length = len(val)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    filename = f\"report_{message.from_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx\"
    wb.save(filename)

    try:
        await message.answer_document(open(filename, 'rb'))
    except Exception as e:
        await message.answer('Ошибка при отправке файла: ' + str(e))
    finally:
        try:
            os.remove(filename)
        except:
            pass

if __name__ == '__main__':
    print('Bot started (milk_bot)')
    executor.start_polling(dp, skip_updates=True)
